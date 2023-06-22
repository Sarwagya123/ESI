from tkinter import *
import tkinter.font as font
from tkinter import messagebox
from tkinter import filedialog
from openpyxl import load_workbook
import datetime
import pandas as pd
import time
import os
import shutil

def main(dest_path):

    def lab(dest_path):
        workbook = load_workbook(filename=dest_path+"Lab.xlsx", data_only=True)
        sheet = workbook.active
        sheet = workbook["Sheet 1"]
        sum=0
        for i in range(14, 47):
            s=str(i)
            if(sheet["G"+s].value is None):
                continue
            else:
                sum+=(float(sheet["F"+s].value)*float(sheet["G"+s].value))
        return sum
    
    def med(dest_path, dates):
        sum=0
        for i in dates:
            workbook = load_workbook(filename=dest_path+i+" Pharmacy.xlsx", data_only=True)
            sheet = workbook.active
            f=False
            for j in range(1, 7):
                sheet = workbook["Sheet"+str(j)]
                for i in range(14,39):
                    s=str(i)
                    if(sheet["B"+s].value is None):
                        f=True
                        break
                    else:
                        sum+=(float(sheet["F"+s].value)*float(sheet["G"+s].value))
                if(f):
                    break      
        return sum

    f=open(dest_path+"patient_details.txt", 'r')
    l=f.readlines()
    doa=l[5].strip()
    dod=l[6].strip()
        #Parse the dates from strings into datetime objects
    date1 = time.mktime(time.strptime(doa, "%d-%m-%Y"))
    date2 = time.mktime(time.strptime(dod, "%d-%m-%Y"))
    
    # No. of days the patient is staying
    nod = int((date2-date1) / 86400) + 1

    # storing the dates the patient was there in the hospital
    test_date = datetime.datetime.strptime(doa, "%d-%m-%Y")
    dates = pd.date_range(test_date, periods=nod)
    dates=map(str, dates)
    dates=[i[0:10] for i in dates]
    for i in range(0,len(dates)):
        year=""
        month=""
        day=""
        for j in range(0, len(dates[i])):
            if(j<=3):
                year=year+dates[i][j]
            elif(j>4 and j<=6):
                month=month+dates[i][j]
            elif(j>7):
                day=day+dates[i][j]
        dates[i]=day+"-"+month+"-"+year

    med_total=med(dest_path, dates)
    print(med_total)
    lab_total=lab(dest_path)

    shutil.copy("res/Main Bill.xlsx", dest_path+"Main Bill.xlsx")
    workbook=load_workbook(filename=dest_path+"Main Bill.xlsx")
    sheet=workbook["Main Bill"]

    sheet["H7"]=dod
    sheet["D11"]=l[0]+" ; "+l[4]+"/"+l[1]
    sheet["D13"]=l[7]
    sheet["D14"]=l[2]
    sheet["D16"]=l[3]
    sheet["D17"]=doa
    sheet["H17"]=dod
    sheet["F26"]=med_total
    sheet["G26"]=1
    sheet["F27"]=lab_total
    sheet["G27"]=1
    workbook.save(filename=dest_path+"Main Bill.xlsx")

    sheet = workbook["Essentiality Certificate"]
    sheet["D8"]=l[0]+" ; "+l[4]+"/"+l[1]
    sheet["D9"]=l[7]
    sheet["D12"]=doa
    sheet["G12"]=dod
    workbook.save(filename=dest_path+"Main Bill.xlsx")

    sheet = workbook["Letter"]
    sheet["D14"]=dod
    sheet["D15"]=l[0]
    sheet["D16"]=l[7]
    workbook.save(filename=dest_path+"Main Bill.xlsx")

    os.startfile(dest_path+"Main Bill.xlsx")

