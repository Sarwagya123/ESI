from tkinter import *
import tkinter.font as font
from tkinter import messagebox
from openpyxl import load_workbook
from ttkwidgets.autocomplete import *
import shutil
import os
import pharmacy
import datetime
import pandas as pd
import time

def main(dest_path, directory, doa, dod, insurance):

     #Parse the dates from strings into datetime objects
    date1 = time.mktime(time.strptime(doa, "%d-%m-%Y"))
    date2 = time.mktime(time.strptime(dod, "%d-%m-%Y"))
    
    # No. of days the patient is staying
    nod = int((date2-date1) / 86400) + 1

    # storing the dates the patient was there in the hospital
    test_date = datetime.datetime.strptime(doa, "%d-%m-%Y")
    dates = pd.date_range(test_date, periods=nod)
    dates=map(str, dates)
    dates=[i[0:10] for i in dates]
    for i in range(0,len(dates)):
        year=""
        month=""
        day=""
        for j in range(0, len(dates[i])):
            if(j<=3):
                year=year+dates[i][j]
            elif(j>4 and j<=6):
                month=month+dates[i][j]
            elif(j>7):
                day=day+dates[i][j]
        dates[i]=day+"-"+month+"-"+year


    #making the required Pharmacy files
    src_path="res/Pharmacy.xlsx"
    for i in range(0,nod):
        print(i)
        shutil.copy(src_path, dest_path+"Pharmacy.xlsx")
        new_name=dates[i]+" Pharmacy.xlsx"
        os.rename(dest_path+"Pharmacy.xlsx", dest_path+new_name)
        workbook = load_workbook(filename=dest_path+new_name)
        sheet = workbook.active
        for j in range(1, 7):
            sheet = workbook["Sheet"+str(j)]
            sheet["E7"] = directory+"(ESIC/S.H.R.C.)"
            sheet["B11"] = "Bill no.ESI//"+insurance
            sheet["H11"] = dates[i]
            workbook.save(filename=dest_path+new_name)

    #making the required Lab File
    src_path="res/Lab.xlsx"
    shutil.copy(src_path, dest_path+"Lab.xlsx")
    workbook = load_workbook(filename=dest_path+"Lab.xlsx")
    sheet = workbook.active
    f=open(dest_path+"patient_details.txt", 'r')
    l=f.readlines()
    sheet["D10"] = l[3]
    sheet["D9"]=l[0]+" ; "+l[4]+"/"+l[1]
    sheet["H7"] = dod
    workbook.save(filename=dest_path+"Lab.xlsx")

    pharmacy.main(dates, dest_path)