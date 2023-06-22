from tkinter import *
import tkinter.font as font
from tkinter import messagebox
from tkinter import filedialog
from openpyxl import load_workbook
import datetime
import pandas as pd
import time
import os
import shutil

def main(dest_path):
    f=open(dest_path+"patient_details.txt", 'r')
    l=f.readlines()
    doa=l[5].strip()
    dod=l[6].strip()
        #Parse the dates from strings into datetime objects
    date1 = time.mktime(time.strptime(doa, "%d-%m-%Y"))
    date2 = time.mktime(time.strptime(dod, "%d-%m-%Y"))
    
    # No. of days the patient is staying
    nod = int((date2-date1) / 86400) + 1

    # storing the dates the patient was there in the hospital
    test_date = datetime.datetime.strptime(doa, "%d-%m-%Y")
    dates = pd.date_range(test_date, periods=nod)
    dates=map(str, dates)
    dates=[i[0:10] for i in dates]
    for i in range(0,len(dates)):
        year=""
        month=""
        day=""
        for j in range(0, len(dates[i])):
            if(j<=3):
                year=year+dates[i][j]
            elif(j>4 and j<=6):
                month=month+dates[i][j]
            elif(j>7):
                day=day+dates[i][j]
        dates[i]=day+"-"+month+"-"+year

    #making the required Challan files
    src_path="res/Challan.xlsx"
    for i in range(0,nod):
        print(i)
        shutil.copy(src_path, dest_path+"Challan.xlsx")
        new_name=dates[i]+" Challan.xlsx"
        os.rename(dest_path+"Challan.xlsx", dest_path+new_name)
        workbook = load_workbook(filename=dest_path+new_name)
        sheet = workbook.active
        for j in range(1, 7):
            sheet = workbook["Sheet"+str(j)]
            sheet["E7"] = l[0]+"(ESIC/S.H.R.C.)"
            sheet["H10"] = dates[i]
            workbook.save(filename=dest_path+new_name)

    for k in dates:
        workbook1 = load_workbook(filename=dest_path+k+" Pharmacy.xlsx", data_only=True)
        workbook2 = load_workbook(filename=dest_path+k+" Challan.xlsx")
        sheet1 = workbook1.active
        sheet2 = workbook2.active
        f=False
        for j in range(1, 7):
            sheet1 = workbook1["Sheet"+str(j)]
            sheet2 = workbook2["Sheet"+str(j)]
            for i in range(14,39):
                s=str(i)
                if(sheet1["B"+s].value is None):
                    f=True
                    break
                else:
                    sheet2["C"+str(i-1)]=sheet1["B"+s].value
                    sheet2["D"+str(i-1)]=sheet1["C"+s].value
                    sheet2["E"+str(i-1)]=sheet1["D"+s].value
                    sheet2["F"+str(i-1)]=sheet1["E"+s].value
                    sheet2["G"+str(i-1)]=sheet1["F"+s].value
                    sheet2["H"+str(i-1)]=sheet1["G"+s].value
            if(f):
                break
        
        workbook2.save(filename=dest_path+k+" Challan.xlsx")