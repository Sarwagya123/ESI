from tkinter import *
import tkinter.font as font
from tkinter import messagebox
from openpyxl import load_workbook
from ttkwidgets.autocomplete import *
import re


def main(dates, dest_path):

    #function to get the details automatically based on name
    def name_auto(t, T1, T2, T3, T4, T5, T6, T7, med_details, med_id, med_name):
        name = T2.get().strip()
        T1.delete('1.0', END)
        # T2.delete(1.0, END)
        T3.delete('1.0', END)
        T4.delete('1.0', END)
        T5.delete('1.0', END)
        T6.delete('1.0', END)
        T7.delete('1.0', END)
        if(name in med_name):
            idx=med_name.index(name)
            T1.insert(INSERT, med_id[idx])
            T5.insert(INSERT, med_details[idx][2])
            T4.insert(INSERT, med_details[idx][3])
            T6.insert(INSERT, med_details[idx][4])
            T7.insert(INSERT, med_details[idx][5])
        else:
            messagebox.showinfo("Add Medicine","No such Medicine in Inventory")
            print("Empty Field!")
        print("Auto Name")

    #function to get the details automatically based on id
    def id_auto(t, T1, T2, T3, T4, T5, T6, T7, med_details, med_id, med_name):
        id = int(T1.get("1.0", END).strip())
        # T1.delete(1.0, END)
        T2.delete(0, 'end')
        T3.delete('1.0', END)
        T4.delete('1.0', END)
        T5.delete('1.0', END)
        T6.delete('1.0', END)
        T7.delete('1.0', END)
        if(id in med_id):
            idx=med_id.index(id)
            T2.insert(INSERT, med_name[idx])
            T5.insert(INSERT, med_details[idx][2])
            T4.insert(INSERT, med_details[idx][3])
            T6.insert(INSERT, med_details[idx][4])
            T7.insert(INSERT, med_details[idx][5])
        else:
            messagebox.showinfo("Add Medicine","No such Medicine in Inventory")
            print("Empty Field!")
        print("Auto ID")

    #function to add the medicine to the patients pharmacy bill
    def add(t, T1, T2, T3, T4, T5, T6, T7, menu, med_details, med_id):
        id = T1.get("1.0", END).strip()
        name = T2.get().strip()
        qty = T3.get("1.0", "end-1c")
        manu = T4.get("1.0", "end-1c")
        batch = T5.get("1.0", "end-1c")
        exp = T6.get("1.0", "end-1c")
        mrp = T7.get("1.0", "end-1c")
        date = menu.get()
        if(id=="" or name=="" or qty=="" or manu=="" or batch=="" or exp=="" or mrp=="" or date=="Select Date"):
                messagebox.showinfo("Add Medicine","Empty Field")
                print("Empty Field!")
        elif(not(re.search("^[0-9]{2}-[0-9]{2}-[0-9]{4}$",exp))):
                messagebox.showinfo("Add Medicine","Add Date in DD-MM-YYYY Format")
                print("Date Format")

        else:
            workbook = load_workbook(filename="res/Inventory.xlsx")
            sheet=workbook.active
            rows = get_maximum_rows(sheet_object=sheet)
            for i in range(2, rows+1):
                if(sheet["A"+str(i)].value==name):
                    sheet["A"+str(i)]=name
                    sheet["B"+str(i)]=batch
                    sheet["C"+str(i)]=manu
                    sheet["D"+str(i)]=exp
                    sheet["E"+str(i)]=mrp   
                    workbook.save(filename="res/Inventory.xlsx")
                    break

            workbook = load_workbook(filename=dest_path+date+" Pharmacy.xlsx")
            sheet = workbook.active
            f=False
            for j in range(1, 7):
                sheet = workbook["Sheet"+str(j)]
                for i in range(14,39):
                    s=str(i)
                    if(sheet["B"+s].value is None):
                        f=True
                        sheet["B"+s]=name
                        sheet["C"+s]=batch
                        sheet["D"+s]=manu
                        sheet["E"+s]=exp
                        sheet["F"+s]=float(med_details[med_id.index(int(id))][5])
                        sheet["G"+s]=int(qty)
                        workbook.save(filename=dest_path+date+" Pharmacy.xlsx")
                        print("add")
                        break
                if(f):
                    break

            T1.delete(1.0, END)
            T2.delete(0, 'end')
            T3.delete('1.0', END)
            T4.delete('1.0', END)
            T5.delete('1.0', END)
            T6.delete('1.0', END)
            T7.delete('1.0', END)

    #function to get the maximum filles rows
    def get_maximum_rows(*, sheet_object):
        rows = 0
        for max_row, row in enumerate(sheet_object, 1):
            if not all(col.value is None for col in row):
                rows += 1
        return rows
    
    inventory_path=r'res\\Inventory.xlsx'
    inventory = load_workbook(inventory_path)
    sheet = inventory.active
    max_row=get_maximum_rows(sheet_object= sheet)
    max_col=sheet.max_column
    med_details=[]
    med_id=[]
    med_name=[]

    #reading the inventory    
    sheet = inventory.active
    for i in range(2, max_row + 1):
        l=[i-1]
        med_id.append(i-1)
        for j in range(1, max_col+1):
            cell_obj = sheet.cell(row = i, column = j)
            if(j==1):
                med_name.append(str(cell_obj.value))
            if(j==4):
                if(isinstance(cell_obj.value, str)):
                    s=cell_obj.value
                elif(len(str(cell_obj.value.month))==1 and len(str(cell_obj.value.day))==1):
                    s="0"+str(cell_obj.value.day)+"-0"+str(cell_obj.value.month)+"-"+str(cell_obj.value.year)
                elif(len(str(cell_obj.value.month))==1 and len(str(cell_obj.value.day))!=1):
                    s=str(cell_obj.value.day)+"-0"+str(cell_obj.value.month)+"-"+str(cell_obj.value.year)
                elif(len(str(cell_obj.value.month))!=1 and len(str(cell_obj.value.day))==1):
                    s="0"+str(cell_obj.value.day)+"-"+str(cell_obj.value.month)+"-"+str(cell_obj.value.year)
                else:
                    s=str(cell_obj.value.day)+"-"+str(cell_obj.value.month)+"-"+str(cell_obj.value.year)
                l.append(s)
                continue
            l.append(str(cell_obj.value))
        med_details.append(l)



    #GUI for adding medicine in the patients pharmacy bill
    window = Tk()

    lbl = Label(window, text="Add Medicines", fg='yellow', bg='blue', font=("Times New Roman", 28))
    lbl.place(x=135, y=15)

    lbl = Label(window, text="Name", fg='white', bg='blue', font=("Times New Roman", 20))
    lbl.place(x=20, y=80)

    T2 = AutocompleteCombobox(window, width=30, completevalues=med_name)
    T2.place(x=95,y=88)

    lbl = Label(window, text="ID", fg='white', bg='blue', font=("Times New Roman", 20))
    lbl.place(x=340, y=80)

    T1 = Text(window, bg='white', fg='black', height=1.4, width=10, padx=1, pady=1)
    T1.place(x=375,y=88)

    lbl = Label(window, text="Manufacturer", fg='white', bg='blue', font=("Times New Roman", 20))
    lbl.place(x=20, y=130)

    T4 = Text(window, bg='white', fg='black', height=1.4, width=22, padx=1, pady=1)
    T4.place(x=180,y=138)

    lbl = Label(window, text="Qty", fg='white', bg='blue', font=("Times New Roman", 20))
    lbl.place(x=355, y=130)

    T3 = Text(window, bg='white', fg='black', height=1.4, width=10, padx=1, pady=1)
    T3.place(x=405,y=138)

    lbl = Label(window, text="Batch", fg='white', bg='blue', font=("Times New Roman", 20))
    lbl.place(x=20, y=183)

    T5 = Text(window, bg='white', fg='black', height=1.4, width=15, padx=1, pady=1)
    T5.place(x=95,y=190)

    lbl = Label(window, text="Expiry", fg='white', bg='blue', font=("Times New Roman", 20))
    lbl.place(x=240, y=183)

    T6 = Text(window, bg='white', fg='black', height=1.4, width=15, padx=1, pady=1)
    T6.place(x=320,y=190)

    lbl = Label(window, text="MRP", fg='white', bg='blue', font=("Times New Roman", 20))
    lbl.place(x=20, y=235)

    T7 = Text(window, bg='white', fg='black', height=1.4, width=15, padx=1, pady=1)
    T7.place(x=85,y=242)

    myFont2 = font.Font(size=5)

    btn1 = Button(window, text="A", fg='black', command=lambda t="Auto Name Button Clicked": name_auto(t, T1, T2, T3, T4, T5, T6, T7, med_details, med_id, med_name))
    btn1['font'] = myFont2
    btn1.place(x=300, y=85)

    btn2 = Button(window, text="A", fg='black', command=lambda t="Auto ID Button Clicked": id_auto(t, T1, T2, T3, T4, T5, T6, T7, med_details, med_id, med_name))
    btn2['font'] = myFont2
    btn2.place(x=460, y=85)

    # declaring the common font size for all the buttons
    myFont = font.Font(size=15)

    #Set the Menu initially
    menu= StringVar(window)
    menu.set("Select Date")

    #Create a dropdown Menu
    drop= OptionMenu(window, menu, *dates)
    drop.place(x=355, y=235)

    btn = Button(window, text="Add", fg='black', command=lambda t="Add Patient Button Clicked": add(t, T1, T2, T3, T4, T5, T6, T7, menu, med_details, med_id))
    btn['font'] = myFont
    btn.place(x=210, y=300)


    window.title('Add Medicines')
    window.geometry("500x360")
    window.configure(bg='blue')
    window.mainloop()