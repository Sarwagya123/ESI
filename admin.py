import shutil
from tkinter import *
import tkinter.font as font
from tkinter import messagebox
from tkinter import filedialog
from ttkwidgets.autocomplete import *
from openpyxl import load_workbook
import add_pat
import edit_pat
import see_pat
import re
import os

def main(n):

    #edit patient details
    def edit(t, T5 , T2, menu1, menu2, window1):
        name=T2.get().strip()
        indoor=T5.get("1.0", "end-1c").strip()
        f=open(r'res\\extras.txt', 'r')
        l=f.readlines()
        if(len(l)==0 or l[0]==""):
            messagebox.showinfo("Edit Patient","First Add a Patient")
            print("See Patient Error")
            f.close()
        
        else:
            parent_dir=l[0]
            folder_name=name+"-"+indoor
            year=menu1.get()
            month=menu2.get()
            if(year=="Select Year" or month=="Select Month" or name=="" or indoor==""):
                messagebox.showinfo("Edit Patient","Please fill up the details")
                print("No Patient Error")
                T2.delete(0, END)
                T5.delete('1.0', END)

            l_year=os.listdir(l[0])

            if(year in l_year):
                l1=os.listdir(parent_dir+year+"/")
                if(month in l1):
                    l2=os.listdir(parent_dir+year+"/"+month+"/")
                    parent_dir=parent_dir+year+"/"+month+"/"
                    if(folder_name in l2):
                        window1.destroy()
                        edit_pat.main(parent_dir+folder_name+"/")
                    else:
                        messagebox.showinfo("Edit Patient","No Patient Found")
                        print("No Paitent Error")
                        T5.delete('1.0', END)
                        T2.delete(0, END)
                else:
                    messagebox.showinfo("Edit Patient","No Month Found")
                    print("No Month Error")
                    T5.delete('1.0', END)
                    T2.delete(0, END)
                    
            else:
                messagebox.showinfo("Edit Patient","No Year Found")
                print("No Year Error")
                T2.delete(0, END)
                T5.delete('1.0', END)

    #function to get the maximum filles rows
    def get_maximum_rows(*, sheet_object):
        rows = 0
        for max_row, row in enumerate(sheet_object, 1):
            if not all(col.value is None for col in row):
                rows += 1
        return rows
    
    #function to add a new patient
    def add_patient(t):
        f=open(r'res\\extras.txt', 'r')
        l=f.readlines()
        if(len(l)==0 or l[0]==""):
            dirpath=filedialog.askdirectory(title="Select Master Folder")
            print(dirpath)
            f.close()
            f=open(r'res\\extras.txt', 'a')
            f.write(dirpath+"/")
            l.append(dirpath+"/")
            f.close()
        add_pat.main(l[0])
        print("add patient")

    #Function to update the inventory
    def update(t, window1, T1, T2, T4, T5, T6):
        inventory_path=r'res\\Inventory.xlsx'
        inventory = load_workbook(inventory_path)
        sheet = inventory.active
        max_row=get_maximum_rows(sheet_object= sheet)
        max_col=sheet.max_column

        name = T2.get("1.0", "end-1c").strip()
        mrp=int(T1.get("1.0", "end-1c").strip())
        manu=T4.get("1.0", "end-1c").strip()
        batch=T5.get("1.0", "end-1c").strip()
        exp=T6.get("1.0", "end-1c").strip()

        if(name=="" or mrp=="" or manu=="" or batch=="" or exp==""):
            messagebox.showinfo("Update Inventory","Empty Field")
            print("Empty Field!")
        elif(not(re.search("^[0-9]{2}-[0-9]{2}-[0-9]{4}$",exp) and re.search("^[0-9]{2}-[0-9]{2}-[0-9]{4}$",exp))):
            messagebox.showinfo("Update Inventory","Add Date in DD-MM-YYYY Format")
            print("Date Format")
        else:
            templ=[name, batch ,manu, exp, mrp]
        
            T1.delete(1.0, "end-1c")
            T2.delete(1.0, "end-1c")
            T4.delete(1.0, "end-1c")
            T5.delete(1.0, "end-1c")
            T6.delete(1.0, "end-1c")

            for i in range(1,max_col+1):
                sheet.cell(row=max_row+1, column=i).value=templ[i-1]
            max_row=max_row+1

            inventory.save(filename="res/Inventory.xlsx")
            window1.destroy()

    def gen_files(t, menu1, menu2, window1, dir):
        year=menu1.get()
        month=menu2.get()
        l_month=os.listdir(dir+year+"/")
        if(month not in l_month):
            messagebox.showinfo("Generate Month Files","No Month Found")
            print("No Month Error")
        else:
            src_path="res/Month Files.xlsx"
            dest_path=dir+year+"/"+month+"/"

            l=os.listdir(dir+year+"/"+month+"/")
            patient_details=[]
            for i in l:
                if(i==month+" Month File.xlsx"):
                    continue
                path=dest_path+i+"/"
                f=open(path+"patient_details.txt",'r')
                lf=f.readlines()
                workbook = load_workbook(filename=path+"Main Bill.xlsx")
                sheet = workbook.active
                sheet = workbook["Main Bill"]
                sum=0.
                for j in range(24,42):
                    if(sheet["F"+str(j)].value is None or sheet["G"+str(j)].value is None):
                        continue
                    else:
                        sum += float(sheet["F"+str(j)].value)*float(sheet["G"+str(j)].value)

                patient_details.append([lf[0], lf[7], lf[5], lf[6], sum, sheet["D18"].value, sheet["D15"].value ])

            if(os.path.isfile(dest_path+month+" Month File.xlsx")):
                os.remove(dest_path+month+" Month File.xlsx")
            shutil.copy(src_path, dest_path+"Month Files.xlsx")
            new_name=month+" Month File.xlsx"
            os.rename(dest_path+"Month Files.xlsx", dest_path+new_name)

            workbook = load_workbook(filename=dest_path+month+" Month File.xlsx")
            sheet = workbook.active
            sheet = workbook["Sheet2"]
            total =0.
            for j in range(len(patient_details)):
                sl=j+1
                s=str(j+11)
                sheet["A"+s]=sl
                sheet["B"+s]=patient_details[j][0]
                sheet["D"+s]=patient_details[j][1]
                sheet["F"+s]=patient_details[j][2]
                sheet["G"+s]=patient_details[j][3]
                sheet["H"+s]=patient_details[j][4]
                total+=patient_details[j][4]
                sheet["J"+s]=patient_details[j][5]
                sheet["E"+s]=patient_details[j][6]
            # for i in range(11,27):
            #     s=str(i)

            sheet = workbook["Sheet1"]
            sheet['B21']="Sub: - Submission of Medical treatment Bill of ESI Patients for the month of "+str(month)
            sheet['B25']="       Please find herewith our Bill No. ……………………. to ………………… amounting to Rs. "+str(total)+". only for the month of "+str(month)
            workbook.save(filename=dest_path+month+" Month File.xlsx")
            os.startfile(dest_path+month+" Month File.xlsx")
            window1.destroy()
            print("Print")

    def gen_month_file(t):
        f=open(r'res\\extras.txt', 'r')
        l=f.readlines()
        if(len(l)==0 or l[0]==""):
            messagebox.showinfo("Generate Monthly Files","First Add a Patient")
            print("Generate Monthly Files Error")
            f.close()


        else:
            l_year=os.listdir(l[0])
            if(len(l_year)==0 or l_year[0]==""):
                messagebox.showinfo("Generate Monthly Files","First Add a Patient")
                print("Generate Monthly Files Error")
                f.close()
            else:
                window1 = Tk()

                #Set the Menu initially
                menu1= StringVar(window1)
                menu1.set("Select Year")

                #Create a dropdown Menu
                drop1= OptionMenu(window1, menu1, *l_year)
                drop1.place(x=80,y=30)

                #Set the Menu initially
                menu2= StringVar(window1)
                menu2.set("Select Month")
                #Create a dropdown Menu
                drop2= OptionMenu(window1, menu2, *["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"])
                drop2.place(x=80,y=80)

                # See button
                btn = Button(window1, text="Generate", fg='black', command=lambda t="Generate Month Files Button Clicked": gen_files(t, menu1, menu2, window1, l[0]))
                btn['font'] = myFont
                btn.place(x=70, y=135)


                window1.title('Generate Month Files')
                window1.geometry("250x200")
                window1.configure(bg='blue')
                window1.mainloop()
                print("Generate month files")


    #function to update the inventory
    def update_inventory(t):

        #GUI for updating inventory
        window1 =Tk()
        lbl = Label(window1, text="Update Inventory", fg='yellow', bg='blue', font=("Times New Roman", 28))
        lbl.place(x=135, y=15)

        lbl = Label(window1, text="Name", fg='white', bg='blue', font=("Times New Roman", 20))
        lbl.place(x=20, y=80)

        T2 = Text(window1, bg='white', fg='black', height=1.4, width=20, padx=1, pady=1)
        T2.place(x=95,y=88)

        lbl = Label(window1, text="MRP", fg='white', bg='blue', font=("Times New Roman", 20))
        lbl.place(x=315, y=80)

        T1 = Text(window1, bg='white', fg='black', height=1.4, width=10, padx=1, pady=1)
        T1.place(x=388,y=88)

        lbl = Label(window1, text="Manufacturer", fg='white', bg='blue', font=("Times New Roman", 20))
        lbl.place(x=20, y=130)

        T4 = Text(window1, bg='white', fg='black', height=1.4, width=30, padx=1, pady=1)
        T4.place(x=180,y=138)

        lbl = Label(window1, text="Batch", fg='white', bg='blue', font=("Times New Roman", 20))
        lbl.place(x=20, y=183)

        T5 = Text(window1, bg='white', fg='black', height=1.4, width=15, padx=1, pady=1)
        T5.place(x=95,y=190)

        lbl = Label(window1, text="Expiry", fg='white', bg='blue', font=("Times New Roman", 20))
        lbl.place(x=240, y=183)

        T6 = Text(window1, bg='white', fg='black', height=1.4, width=15, padx=1, pady=1)
        T6.place(x=320,y=190)


        # declaring the common font size for all the buttons
        myFont = font.Font(size=15)

        btn = Button(window1, text="Update", fg='black', command=lambda t="Add Patient Button Clicked": update(t,window1, T1, T2, T4, T5,T6))
        btn['font'] = myFont
        btn.place(x=190, y=260)

        window1.title('Update Inventory')
        window1.geometry("500x350")
        window1.configure(bg='blue')
        window1.mainloop()

    
    def name_auto(t, T2, T5, l_pat_dict):
        name=T2.get().strip()
        if(name in l_pat_dict.keys()):
            T5.insert(INSERT, l_pat_dict[name])
        else:
            messagebox.showinfo("Edit Patient","No Patient Found")
            print("Empty Field!")
        print("Auto Name")

    def edit_patients_see(t, menu1, menu2, window1):

        f=open(r'res\\extras.txt', 'r')
        l=f.readlines()
        l_pat_dict=dict()
        l_pat_name=[]
        if(len(l)==0 or l[0]==""):
            messagebox.showinfo("Edit Patient","First Add a Patient")
            print("Edit Patient Error")
        
        else:
            parent_dir=l[0]
            year=menu1.get()
            month=menu2.get()
            if(year=="Select Year" or month=="Select Month"):
                messagebox.showinfo("Edit Patient","Please fill up the details")
                print("No Patient Error")
                return
            else:
                l_year=os.listdir(l[0])
                if(year in l_year):
                    l1=os.listdir(parent_dir+year+"/")
                    if(month not in l1):
                        messagebox.showinfo("Edit Patient","No Month Found")
                        print("No Patient Error")
                        return
                    else:
                        window1.destroy()
                        l_pat=os.listdir(parent_dir+year+"/"+month+"/")
                        for pat in l_pat:
                            pat=pat.split("-")
                            l_pat_name.append(pat[0])
                            l_pat_dict[pat[0]]=pat[1]
        

        window2 = Tk()
        lbl = Label(window2, text="Name", fg='white', bg='blue', font=("Times New Roman", 20))
        lbl.place(x=20, y=20)

        T2 = AutocompleteCombobox(window2, width=30, completevalues=l_pat_name)
        T2.place(x=95,y=28)

        # T2 = Text(window2, bg='white', fg='black', height=1.4, width=30, padx=1, pady=1)
        # T2.place(x=95,y=28)

        lbl = Label(window2, text="Indoor Reg. No.", fg='white', bg='blue', font=("Times New Roman", 20))
        lbl.place(x=20, y=73)

        T5 = Text(window2, bg='white', fg='black', height=1.4, width=15, padx=1, pady=1)
        T5.place(x=200,y=80)

        # See button
        btn = Button(window2, text="Ok", fg='black', command=lambda t="See Patient Button Clicked": edit(t, T5, T2, menu1, menu2, window2))
        btn['font'] = myFont
        btn.place(x=160, y=120)

        myFont2 = font.Font(size=5)

        btn1 = Button(window2, text="A", fg='black', command=lambda t="Auto Name Button Clicked": name_auto(t, T2, T5, l_pat_dict))
        btn1['font'] = myFont2
        btn1.place(x=310, y=25)

        window2.title('Edit Patient')
        window2.geometry("350x180")
        window2.configure(bg='blue')
        window2.mainloop()
    
    #function to edit the patient details
    def edit_patients(t):
        f=open(r'res\\extras.txt', 'r')
        l=f.readlines()
        if(len(l)==0 or l[0]==""):
            messagebox.showinfo("Edit Patient","First Add a Patient")
            print("Edit Patient Error")
            f.close()


        else:
            l_year=os.listdir(l[0])
            if(len(l_year)==0 or l_year[0]==""):
                messagebox.showinfo("Edit Patient","First Add a Patient")
                print("Edit Patient Error")
                f.close()
            else:

                window1 = Tk()

                # declaring the common font size for all the buttons
                myFont = font.Font(size=15)

                #Set the Menu initially
                menu1= StringVar(window1)
                menu1.set("Select Year")

                #Create a dropdown Menu
                drop1= OptionMenu(window1, menu1, *l_year)
                drop1.place(x=30,y=20)

                #Set the Menu initially
                menu2= StringVar(window1)
                menu2.set("Select Month")
                #Create a dropdown Menu
                drop2= OptionMenu(window1, menu2, *["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"])
                drop2.place(x=200,y=20)

                # See button
                btn = Button(window1, text="Ok", fg='black', command=lambda t="See Patient Button Clicked": edit_patients_see(t, menu1, menu2, window1))
                btn['font'] = myFont
                btn.place(x=160, y=70)


                window1.title('Edit Patient')
                window1.geometry("350x120")
                window1.configure(bg='blue')
                window1.mainloop()
                print("edit patients")


    #function to See patient details
    def see_patients(t):
        see_pat.main()


    #Admin GUI
    window = Tk()

    lbl = Label(window, text="Welcome User", fg='yellow', bg='blue', font=("Times New Roman", 30))
    lbl.place(x=35, y=20)

    # declaring the common font size for all the buttons
    myFont = font.Font(size=15)

    # Inventory button
    btn = Button(window, text="Update Inventory", fg='black', command=lambda t="INVENTORY Button Clicked": update_inventory(t))
    btn['font'] = myFont
    btn.place(x=80, y=105)

    # Generate Monthy Files button
    btn = Button(window, text="Generate Monthly Files", fg='black', command=lambda t="GENERATE MONTHLY FILES Button Clicked": gen_month_file(t))
    btn['font'] = myFont
    btn.place(x=50, y=195)

    # Add Patient button
    btn = Button(window, text="Add Patient", fg='black', command=lambda t="ADD PATIENT Button Clicked": add_patient(t))
    btn['font'] = myFont
    btn.place(x=100, y=285)

    # Edit Patient button
    btn = Button(window, text="Edit Patient", fg='black', command=lambda t="EDIT PATIENT Button Clicked": edit_patients(t))
    btn['font'] = myFont
    btn.place(x=100, y=375)

    # See Patient button
    btn = Button(window, text="See Patient", fg='black', command=lambda t="SEE PATIENTS Button Clicked": see_patients(t))
    btn['font'] = myFont
    btn.place(x=100, y=465)

    window.title('Admin')
    window.geometry("300x550")
    window.configure(bg='blue')
    window.mainloop()